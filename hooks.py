import base64
import calendar
import glob
import logging
import os
import time
from datetime import date, datetime, timedelta

import requests
from openpyxl import load_workbook
from psycopg2.errors import SerializationFailure

from odoo.exceptions import UserError
from odoo.fields import Command

_logger = logging.getLogger(__name__)

# New xml_ids created by this import (e.g. for a journal/asset row that has
# no existing match) are registered under this addon's own name, regardless
# of which client module's data actually triggered the import.
MODULE = os.path.basename(os.path.dirname(__file__))

JOURNAL_TYPES = {"Bank": "bank", "Kas": "cash", "Lain-lain": "general"}
ASSET_METHODS = {"Garis Lurus": "linear"}
ASSET_PERIODS = {"Bulan": "1", "Tahun": "12"}
ACCOUNT_TYPES = {
    "piutang": "asset_receivable",
    "bank dan tunai": "asset_cash",
    "aktiva lancar": "asset_current",
    "aktiva tidak lancar": "asset_non_current",
    "prabayar": "asset_prepayments",
    "aktiva tetap": "asset_fixed",
    "utang": "liability_payable",
    "kartu kredit": "liability_credit_card",
    "pasiva terkini": "liability_current",
    "hutang tidak lancar": "liability_non_current",
    "ekuitas": "equity",
    "penghasilan tahun terkini": "equity_unaffected",
    "penghasilan": "income",
    "penghasilan lainnya": "income_other",
    "pengeluaran": "expense",
    "pengeluaran lainnya": "expense_other",
    "penyusutan": "expense_depreciation",
    "biaya pendapatan": "expense_direct_cost",
    "off-balance sheet": "off_balance",
}
ANALYTIC_PLAN_APPLICABILITIES = {
    "optional": "optional",
    "opsional": "optional",
    "mandatory": "mandatory",
    "wajib": "mandatory",
    "unavailable": "unavailable",
    "tidak tersedia": "unavailable",
}

REPORT_RELATED_REFS = {
    "account_reports.balance_sheet": {
        "action": "account_reports.action_account_report_bs",
        "menu": "account_reports.menu_action_account_report_balance_sheet",
    },
    "account_reports.profit_and_loss": {
        "action": "account_reports.action_account_report_pl",
        "menu": "account_reports.menu_action_account_report_profit_and_loss",
    },
    "account_reports.cash_flow_report": {
        "action": "account_reports.action_account_report_cs",
        "menu": "account_reports.menu_action_account_report_cash_flow",
    },
    "account_reports.executive_summary": {
        "action": "account_reports.action_account_report_exec_summary",
        "menu": "account_reports.menu_action_account_report_exec_summary",
    },
    "account_reports.trial_balance_report": {
        "action": "account_reports.action_account_report_coa",
        "menu": "account_reports.menu_action_account_report_coa",
    },
    "account_reports.general_ledger_report": {
        "action": "account_reports.action_account_report_general_ledger",
        "menu": "account_reports.menu_action_account_report_general_ledger",
    },
    "account_reports.partner_ledger_report": {
        "action": "account_reports.action_account_report_partner_ledger",
        "menu": "account_reports.menu_action_account_report_partner_ledger",
    },
    "account_reports.aged_receivable_report": {
        "action": "account_reports.action_account_report_ar",
        "menu": "account_reports.menu_action_account_report_aged_receivable",
    },
    "account_reports.aged_payable_report": {
        "action": "account_reports.action_account_report_ap",
        "menu": "account_reports.menu_action_account_report_aged_payable",
    },
}

# Used when the "company" sheet's external_report_layout row is blank -
# every client so far uses the same layout, but the sheet can still override it.
DEFAULT_EXTERNAL_REPORT_LAYOUT = "web.external_layout_folder"

# "company" sheet key -> res.config.settings field toggled through it. Both
# are applied via res.config.settings.execute(), not company.write(), since
# that's what actually applies the group implication / triggers the module
# install - same as checking the box in Settings and clicking Save.
ACCOUNTING_FEATURE_FIELDS = {
    "analytic_accounting": "group_analytic_accounting",
    "budget_management": "module_account_budget",  # installs account_budget
}

_TRUE_VALUES = {"true", "ya", "yes", "1", "aktif"}
_FALSE_VALUES = {"false", "tidak", "no", "0", "nonaktif", "non-aktif"}


def _parse_bool(value):
    """Parse a yes/no "company" sheet cell. None/blank means "not specified,
    don't touch" - unlike the other company fields, a boolean's off state is
    a real value someone might want to set, so this needs three outcomes
    (True/False/None) instead of just skip-if-blank.
    """
    if isinstance(value, bool):
        return value
    if value in (None, ""):
        return None
    text = str(value).strip().lower()
    if text in _TRUE_VALUES:
        return True
    if text in _FALSE_VALUES:
        return False
    _logger.warning('Unrecognized yes/no value %r in the "company" sheet - ignoring.', value)
    return None

# Set to True to auto-validate imported assets (generates the depreciation
# board and posts its moves). False leaves them as draft for manual review.
VALIDATE_IMPORTED_ASSETS = False

# Set to True to auto-post imported opening vendor/customer moves.
# False leaves them as draft for manual review.
POST_OPENING_MOVES = False


def find_data_file(data_dir):
    """Path to the single .xlsx under data_dir, or None if there isn't one -
    the data import is optional seed data, not a structural part of a client
    module, so a missing file just skips the import instead of blocking
    install. More than one file is treated as a real misconfiguration
    (which one would we import?) and still raises.
    """
    matches = glob.glob(os.path.join(data_dir, "*.xlsx"))
    if len(matches) > 1:
        raise ValueError(f"Expected at most one .xlsx file in {data_dir}, found {len(matches)}")
    return matches[0] if matches else None


def import_bundled_data(env, module_dir):
    """Convenience entry point for a client module's post_init_hook: find
    the single .xlsx under <module_dir>/data and run the import, or skip
    with a warning if there isn't one.
    """
    data_dir = os.path.join(module_dir, "data")
    path = find_data_file(data_dir)
    if not path:
        _logger.warning("No .xlsx master data file found under %s - skipping data import.", data_dir)
        return
    wb = load_workbook(path, read_only=True, data_only=True)
    run_import(env, wb)


_DATE_STRING_FORMATS = ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y")


def _parse_sheet_date(value, context):
    """A date cell normally comes back from openpyxl as a real datetime,
    but if someone typed the date as plain text in Google Sheets instead
    of a proper Date-formatted cell, it comes back as a str instead -
    try a few common formats before giving up, so that crashes with a
    clear message instead of an opaque "'str' object has no attribute
    'date'" AttributeError.
    """
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        text = value.strip()
        for fmt in _DATE_STRING_FORMATS:
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                continue
    raise ValueError(
        f"{context}: nilai tanggal {value!r} tidak dikenali - pastikan cell-nya "
        "berformat Date (bukan Text) di sheet-nya."
    )


def _read_opening_balance_date(wb):
    """Read OPENING_BALANCE_DATE, a key/value row in the "company" sheet
    (falls back to the older "petunjuk" location for workbooks that haven't
    been migrated yet). Ini tanggal cutover opening balance yang sebenarnya
    (mis. "saldo per 30 Juni 2026") - fiscal_year_start di run_import()
    di-derive dari sini (+1 hari), bukan sebaliknya, karena inilah nilai
    yang paling natural diisi orang: tanggal per kapan saldo awal berlaku.
    """
    for sheet in ("company", "petunjuk"):
        if sheet not in wb.sheetnames:
            continue
        for row in wb[sheet].iter_rows(values_only=True):
            if row and row[0] == "OPENING_BALANCE_DATE":
                return _parse_sheet_date(row[1], "OPENING_BALANCE_DATE")
    raise ValueError('"OPENING_BALANCE_DATE" tidak ditemukan di sheet "company" atau "petunjuk"')


def _sheet_rows(wb, sheet, columns):
    if sheet not in wb.sheetnames:
        return
    rows = wb[sheet].iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration:
        return
    if not header:
        return
    # A column entirely absent from the header (not just blank cells) reads
    # as None for every row - same tolerance _move_rows already gives an
    # optional column, so a sheet can omit one it doesn't need at all.
    idx = {col: header.index(col) for col in columns if col in header}
    for row in rows:
        values = {col: (row[idx[col]] if col in idx and idx[col] < len(row) else None) for col in columns}
        if values.get("id"):
            values["id"] = values["id"] if "." in str(values["id"]) else f"{MODULE}.{values['id']}"
            yield values


def _get_or_create(env, model, xml_id, values):
    record = env.ref(xml_id, raise_if_not_found=False)
    if record:
        record.write(values)
        return record
    # env.ref() returns None both when the xml_id was never registered and
    # when its target record was deleted without cleaning up ir.model.data
    # (e.g. manual deletion in the UI) - reuse/repoint a stale entry instead
    # of blindly creating a duplicate (module, name) row.
    record = env[model].create(values)
    module, name = xml_id.split(".", 1)
    data = env["ir.model.data"].search([("module", "=", module), ("name", "=", name)])
    if data:
        data.write({"model": model, "res_id": record.id, "noupdate": True})
    else:
        env["ir.model.data"].create({"module": module, "name": name, "model": model, "res_id": record.id, "noupdate": True})
    return record


def _account_by_code_name(env, code_name):
    if not code_name:
        return env["account.account"]
    code = code_name.split(" ", 1)[0]
    return env["account.account"].with_context(active_test=False).search([("code", "=", code)], limit=1)


def _elapsed_depreciation_periods(acquisition_date, balance_date, method_period):
    """Prorata periods from acquisition_date through balance_date, matching
    Odoo 19 constant_periods prorata calculation convention.
    """
    if not acquisition_date or not balance_date or acquisition_date > balance_date:
        return 0.0

    days_in_acq_month = calendar.monthrange(acquisition_date.year, acquisition_date.month)[1]
    days_in_bal_month = calendar.monthrange(balance_date.year, balance_date.month)[1]

    start_prorata = (days_in_acq_month - acquisition_date.day + 1) / days_in_acq_month
    end_prorata = balance_date.day / days_in_bal_month

    elapsed_months = (
        start_prorata
        + end_prorata
        + (balance_date.year - acquisition_date.year) * 12
        + (balance_date.month - acquisition_date.month - 1)
    )

    if method_period == "12":  # yearly
        return max(0.0, elapsed_months / 12.0)
    return max(0.0, elapsed_months)


def _write_opening_balances(env, totals):
    """totals: dict of account.account recordset -> [opening_debit, opening_credit].
    Writing these fields queues a rebuild of the company's single opening
    balance move, which Odoo refuses once that move is posted - skip
    entirely once the user has reviewed and posted it, instead of crashing.
    """
    opening_move = env.ref("base.main_company").account_opening_move_id
    if opening_move and opening_move.state != "draft":
        return
    for account, (debit, credit) in totals.items():
        values = {}
        if debit:
            values["opening_debit"] = debit
        if credit:
            values["opening_credit"] = credit
        if values:
            account.write(values)


def _partner_by_name(env, name, rank_field):
    if not name:
        return env["res.partner"]
    partner = env["res.partner"].search([("name", "=", name)], limit=1)
    return partner or env["res.partner"].create({"name": name, rank_field: 1})


def _move_rows(wb, sheet, default_date):
    """Group an Odoo-import-style sheet (a header row per move, followed by
    blank-id continuation rows) into one dict per move with a "lines" list.
    """
    rows = wb[sheet].iter_rows(values_only=True)
    header = next(rows)
    idx = {col: header.index(col) for col in header}

    def get(row, col):
        i = idx.get(col)
        return row[i] if i is not None and i < len(row) else None

    move = None
    for row in rows:
        xml_id = get(row, "id")
        if xml_id:
            if move:
                yield move
            xml_id = xml_id if "." in xml_id else f"{MODULE}.{xml_id}"
            raw_date = get(row, "date")
            move = {
                "id": xml_id,
                "ref": get(row, "Reference"),
                "date": _parse_sheet_date(raw_date, "date") if raw_date else default_date,
                "journal": get(row, "journal"),
                "lines": [],
            }
        if move is None:
            continue
        account = get(row, "line_ids/account")
        if account:
            move["lines"].append({
                "account": account,
                "debit": get(row, "line_ids/debit"),
                "credit": get(row, "line_ids/credit"),
                "name": get(row, "line_ids/name"),
                "partner": get(row, "line_ids/partner"),
                "date_maturity": get(row, "line_ids/date_maturity"),
            })
    if move:
        yield move


def _import_opening_move(env, wb, sheet, partner_rank_field, default_date):
    company = env.ref("base.main_company")
    for move_data in _move_rows(wb, sheet, default_date):
        existing = env.ref(move_data["id"], raise_if_not_found=False)
        if existing and existing.state != "draft":
            continue

        line_commands = []
        balance = 0.0
        for line in move_data["lines"]:
            debit, credit = line["debit"] or 0.0, line["credit"] or 0.0
            balance += credit - debit
            vals = {
                "account_id": _account_by_code_name(env, line["account"]).id,
                "debit": debit,
                "credit": credit,
                "name": line["name"],
                "partner_id": _partner_by_name(env, line["partner"], partner_rank_field).id,
            }
            if line["date_maturity"]:
                vals["date_maturity"] = _parse_sheet_date(line["date_maturity"], "date_maturity")
            line_commands.append(Command.create(vals))

        if not company.currency_id.is_zero(balance):
            balancing_account = company.get_unaffected_earnings_account()
            line_commands.append(Command.create({
                "account_id": balancing_account.id,
                "debit": max(balance, 0.0),
                "credit": max(-balance, 0.0),
                "name": "Automatic Balancing Line",
            }))

        values = {
            "ref": move_data["ref"],
            "date": move_data["date"],
            "journal_id": env["account.journal"].search([("name", "=", move_data["journal"])], limit=1).id,
            "move_type": "entry",
            "line_ids": [Command.clear()] + line_commands,
        }
        move = _get_or_create(env, "account.move", move_data["id"], values)
        if POST_OPENING_MOVES and move.state == "draft":
            move.action_post()


def _import_company(env, wb):
    """Read the "company" sheet - key/value rows (also where
    OPENING_BALANCE_DATE lives, see _read_opening_balance_date) - and write
    it onto the single company record, then apply any accounting feature
    toggles found there (see
    ACCOUNTING_FEATURE_FIELDS). Optional sheet, and a blank cell leaves that
    field/toggle untouched rather than clearing it - so a partially-filled
    sheet (e.g. no logo URL yet) never blanks out data set another way.
    """
    if "company" not in wb.sheetnames:
        return
    data = {}
    for row in wb["company"].iter_rows(values_only=True):
        if row and row[0]:
            data[row[0]] = row[1]
    if not data:
        return

    company = env.ref("base.main_company")
    values = {}
    for field in ("name", "street", "street2", "city", "zip", "phone", "email", "website", "report_footer"):
        value = data.get(field)
        if value in (None, ""):
            continue
        # A numeric-looking text field (e.g. zip) can come back as a float
        # if the sheet cell got auto-formatted as a number - avoid writing
        # "55151.0" instead of "55151".
        if isinstance(value, float) and value.is_integer():
            value = int(value)
        values[field] = str(value)

    if data.get("country"):
        country = env["res.country"].search([("name", "=", data["country"])], limit=1)
        values["country_id"] = country.id
        if data.get("state"):
            state = env["res.country.state"].search(
                [("name", "=", data["state"]), ("country_id", "=", country.id)], limit=1
            )
            values["state_id"] = state.id

    layout = env.ref(data.get("external_report_layout") or DEFAULT_EXTERNAL_REPORT_LAYOUT, raise_if_not_found=False)
    if layout:
        values["external_report_layout_id"] = layout.id

    logo_url = data.get("logo")
    if logo_url:
        try:
            response = requests.get(logo_url, timeout=10)
            response.raise_for_status()
            values["logo"] = base64.b64encode(response.content)
        except Exception:
            _logger.warning("Failed to download company logo from %s - keeping existing logo.", logo_url, exc_info=True)

    company.write(values)

    settings_values = {}
    for key, field in ACCOUNTING_FEATURE_FIELDS.items():
        parsed = _parse_bool(data.get(key))
        if parsed is not None:
            settings_values[field] = parsed
    if settings_values:
        _apply_config_settings(env, company, values, settings_values)


def _apply_config_settings(env, company, company_values, settings_values):
    """A module_* setting (e.g. budget_management -> module_account_budget)
    goes through Odoo's live "hot install" path, which takes an exclusive
    lock on ir_cron as a safety check - this can lose a race against a
    concurrently running server's own cron-polling thread and abort with a
    SerializationFailure. That's a textbook case for "just retry the
    transaction" - rolling back also discards the company.write() above
    (same uncommitted transaction), so redo it before each retry.
    """
    for attempt in range(3):
        try:
            env["res.config.settings"].create(settings_values).execute()
            return
        except SerializationFailure:
            if attempt == 2:
                raise
            env.cr.rollback()
            company.write(company_values)
            time.sleep(1)


def _import_res_partner(env, wb):
    columns = ("id", "name", "email", "phone", "is_company", "street", "city", "state", "country_id", "ref")
    for row in _sheet_rows(wb, "res.partner", columns):
        country = env["res.country"]
        if row["country_id"]:
            country = country.search([("name", "=", row["country_id"])], limit=1)
        state = env["res.country.state"]
        if row["state"]:
            state = state.search([("name", "=", row["state"]), ("country_id", "=", country.id)], limit=1)
        values = {
            "name": row["name"],
            "email": row["email"],
            "phone": row["phone"],
            "is_company": bool(row["is_company"]),
            "street": row["street"],
            "city": row["city"],
            "state_id": state.id,
            "country_id": country.id,
            "ref": row["ref"],
        }
        _get_or_create(env, "res.partner", row["id"], values)


def _import_account_analytic_plan(env, wb):
    columns = ("id", "name", "sequence", "default_applicability", "color", "description", "parent_id")
    for row in _sheet_rows(wb, "account.analytic.plan", columns):
        values = {
            "name": row["name"],
        }
        if row["sequence"] is not None:
            values["sequence"] = int(row["sequence"])
        if row["color"] is not None:
            values["color"] = int(row["color"])
        if row["default_applicability"]:
            raw_app = str(row["default_applicability"]).strip().lower()
            if raw_app in ANALYTIC_PLAN_APPLICABILITIES:
                values["default_applicability"] = ANALYTIC_PLAN_APPLICABILITIES[raw_app]
        if row["description"]:
            values["description"] = str(row["description"])
        if row["parent_id"]:
            parent = env.ref(row["parent_id"], raise_if_not_found=False)
            if not parent and "." not in str(row["parent_id"]):
                parent = env.ref(f"{MODULE}.{row['parent_id']}", raise_if_not_found=False)
            if not parent:
                parent = env["account.analytic.plan"].search([("name", "=", row["parent_id"])], limit=1)
            if parent:
                values["parent_id"] = parent.id
        _get_or_create(env, "account.analytic.plan", row["id"], values)


def _import_account_analytic_account(env, wb):
    columns = ("id", "name", "plan_id", "code", "active", "partner_id")
    for row in _sheet_rows(wb, "account.analytic.account", columns):
        plan = None
        if row["plan_id"]:
            plan_str = str(row["plan_id"]).strip()
            plan = env.ref(plan_str, raise_if_not_found=False)
            if not plan and "." not in plan_str:
                plan = env.ref(f"{MODULE}.{plan_str}", raise_if_not_found=False)
            if not plan:
                plan = env["account.analytic.plan"].search([("name", "=", plan_str)], limit=1)
            if not plan:
                plan = env["account.analytic.plan"].search([("complete_name", "=", plan_str)], limit=1)

        if not plan:
            _logger.warning(
                "Rencana Analitik (plan_id) %r tidak ditemukan untuk akun analitik %r (id=%s) - dilewati.",
                row["plan_id"], row["name"], row["id"]
            )
            continue

        values = {
            "name": row["name"],
            "plan_id": plan.id,
        }
        if row["code"]:
            code_val = str(int(row["code"])) if isinstance(row["code"], float) and row["code"].is_integer() else str(row["code"])
            values["code"] = code_val
        if row["active"] is not None:
            parsed_active = _parse_bool(row["active"])
            if parsed_active is not None:
                values["active"] = parsed_active
        if row["partner_id"]:
            partner = env["res.partner"].search([("name", "=", row["partner_id"])], limit=1)
            if partner:
                values["partner_id"] = partner.id
        _get_or_create(env, "account.analytic.account", row["id"], values)


def _import_account_report(env, wb):
    if "account.report" not in wb.sheetnames or "account.report" not in env:
        return
    columns = ("id", "name")
    for row in _sheet_rows(wb, "account.report", columns):
        if not row["name"]:
            continue
        raw_id = str(row["id"]).strip()
        xml_id = raw_id
        if xml_id.startswith(f"{MODULE}.account_reports."):
            xml_id = xml_id[len(f"{MODULE}."):]
        elif xml_id.startswith(f"{MODULE}.account."):
            xml_id = xml_id[len(f"{MODULE}."):]

        report = env.ref(xml_id, raise_if_not_found=False)
        if not report and "." not in xml_id:
            report = env.ref(f"account_reports.{xml_id}", raise_if_not_found=False)
        if not report:
            report = env["account.report"].search([("name", "=", xml_id)], limit=1)

        if not report:
            _logger.warning("Laporan accounting (account.report) %r tidak ditemukan - dilewati.", xml_id)
            continue

        report.write({"name": row["name"]})

        lookup_id = xml_id if "." in xml_id else f"account_reports.{xml_id}"
        refs = REPORT_RELATED_REFS.get(lookup_id) or REPORT_RELATED_REFS.get(xml_id)
        if refs:
            action = env.ref(refs["action"], raise_if_not_found=False)
            if action:
                action.write({"name": row["name"]})
            menu = env.ref(refs["menu"], raise_if_not_found=False)
            if menu:
                menu.write({"name": row["name"]})
        else:
            actions = env["ir.actions.client"].search([("tag", "=", "account_report")])
            for act in actions:
                if str(report.id) in str(act.context):
                    act.write({"name": row["name"]})
                    menus = env["ir.ui.menu"].search([("action", "=", f"ir.actions.client,{act.id}")])
                    if menus:
                        menus.write({"name": row["name"]})


def _import_account_report_line(env, wb):
    if "account.report.line" not in wb.sheetnames or "account.report.line" not in env:
        return
    columns = ("id", "name", "code")
    for row in _sheet_rows(wb, "account.report.line", columns):
        if not row["name"]:
            continue
        raw_id = str(row["id"]).strip()
        xml_id = raw_id
        if xml_id.startswith(f"{MODULE}.account_reports."):
            xml_id = xml_id[len(f"{MODULE}."):]
        elif xml_id.startswith(f"{MODULE}.account."):
            xml_id = xml_id[len(f"{MODULE}."):]

        line = env.ref(xml_id, raise_if_not_found=False)
        if not line and "." not in xml_id:
            line = env.ref(f"account_reports.{xml_id}", raise_if_not_found=False)
        if not line and row.get("code"):
            line = env["account.report.line"].search([("code", "=", str(row["code"]).strip())], limit=1)
        if not line:
            line = env["account.report.line"].search([("name", "=", xml_id)], limit=1)

        if not line:
            _logger.warning("Baris laporan (account.report.line) %r tidak ditemukan - dilewati.", xml_id)
            continue

        line.write({"name": row["name"]})


def _import_kas_bank(env, wb, default_date):
    """Import kas/bank opening balances as account.bank.statement.line records.
    Source:
    1. Direct opening_balance column on sheet "account.journal" (preferred).
    2. Legacy sheet "kas_bank" (fallback if sheet exists and has data).
    """
    company = env.ref("base.main_company")
    transfer_total = 0.0

    # 1. Preferred source: read from sheet "account.journal" column "opening_balance"
    journal_has_balance = False
    if "account.journal" in wb.sheetnames:
        columns = ("id", "name", "type", "code", "opening_balance")
        rows = list(_sheet_rows(wb, "account.journal", columns))
        if any(r.get("opening_balance") not in (None, "") for r in rows):
            journal_has_balance = True

            # Clean up all existing opening statement lines on default_date for bank/cash journals
            all_bank_cash = env["account.journal"].search([
                ("type", "in", ("bank", "cash")),
                ("company_id", "=", company.id),
            ])
            existing_lines = env["account.bank.statement.line"].search([
                ("journal_id", "in", all_bank_cash.ids),
                ("date", "=", default_date),
            ])
            for old in existing_lines:
                if old.is_reconciled:
                    old.line_ids.remove_move_reconcile()
                old.unlink()

            for row in rows:
                raw_bal = row.get("opening_balance")
                if raw_bal in (None, ""):
                    continue
                try:
                    amount = float(raw_bal)
                except (ValueError, TypeError):
                    _logger.warning("Nilai opening_balance %r pada jurnal %r tidak valid - dilewati.", raw_bal, row.get("name"))
                    continue

                journal = env.ref(row["id"], raise_if_not_found=False)
                if not journal and row.get("code"):
                    journal = env["account.journal"].search([
                        ("code", "=", str(row["code"]).strip()),
                        ("company_id", "=", company.id),
                    ], limit=1)
                if not journal and row.get("name"):
                    journal = env["account.journal"].search([
                        ("name", "=", str(row["name"]).strip()),
                        ("company_id", "=", company.id),
                    ], limit=1)

                if not journal:
                    _logger.warning("Jurnal %r tidak ditemukan untuk saldo awal kas/bank - dilewati.", row.get("name"))
                    continue

                transfer_total += amount
                line_xml_id = f"{MODULE}.statement_line_{row['id'].replace('.', '_')}"



                values = {
                    "journal_id": journal.id,
                    "date": default_date,
                    "payment_ref": "Saldo Awal",
                    "amount": amount,
                    "counterpart_account_id": company.transfer_account_id.id,
                }
                _get_or_create(env, "account.bank.statement.line", line_xml_id, values)

    # 2. Fallback source: legacy sheet "kas_bank"
    if not journal_has_balance and "kas_bank" in wb.sheetnames:
        for move_data in _move_rows(wb, "kas_bank", default_date):
            journal = env["account.journal"].search([
                ("name", "=", move_data["journal"]),
                ("company_id", "=", company.id),
            ], limit=1)
            if not journal:
                _logger.warning("Jurnal %r untuk kas_bank tidak ditemukan - dilewati.", move_data["journal"])
                continue

            amount = sum(
                (line["debit"] or 0.0) - (line["credit"] or 0.0)
                for line in move_data["lines"]
                if _account_by_code_name(env, line["account"]) == journal.default_account_id
            )
            transfer_total += amount

            existing = env.ref(move_data["id"], raise_if_not_found=False)
            if existing:
                if existing.is_reconciled:
                    _logger.warning("Bank statement line %r sudah direkonsiliasi - dilewati.", move_data["id"])
                    continue
                existing.unlink()

            values = {
                "journal_id": journal.id,
                "date": move_data["date"],
                "payment_ref": move_data["ref"],
                "amount": amount,
                "counterpart_account_id": company.transfer_account_id.id,
            }
            _get_or_create(env, "account.bank.statement.line", move_data["id"], values)

    _write_opening_balances(env, {
        company.transfer_account_id: [max(transfer_total, 0.0), max(-transfer_total, 0.0)],
    })


def _import_account_account(env, wb):
    columns = ("id", "code", "name", "account_type", "opening_debit", "opening_credit", "active")
    opening_totals = {}
    company = env.ref("base.main_company")
    for row in _sheet_rows(wb, "account.account", columns):
        code = str(row["code"]).strip() if row["code"] is not None else ""
        name = str(row["name"]).strip() if row["name"] is not None else ""
        if not code or not name:
            continue

        raw_type = row.get("account_type")
        account_type = None
        if raw_type:
            raw_str = str(raw_type).strip().lower()
            account_type = ACCOUNT_TYPES.get(raw_str) or (raw_str if raw_str in ACCOUNT_TYPES.values() else None)

        values = {"code": code, "name": name}
        if account_type:
            values["account_type"] = account_type

        if row.get("active") is not None:
            parsed_active = _parse_bool(row["active"])
            if parsed_active is not None:
                values["active"] = parsed_active

        account = False
        if code:
            account = env["account.account"].with_context(active_test=False).search([
                ("code", "=", code),
                ("company_ids", "in", company.id),
            ], limit=1)
            if not account:
                account = env["account.account"].with_context(active_test=False).search([("code", "=", code)], limit=1)
        if not account:
            cand = env.ref(row["id"], raise_if_not_found=False)
            if cand and (not cand.code or cand.code == code):
                account = cand

        if account:
            account.write(values)
            if "." in row["id"]:
                module, xml_name = row["id"].split(".", 1)
            else:
                module, xml_name = MODULE, row["id"]
            data = env["ir.model.data"].search([("module", "=", module), ("name", "=", xml_name)])
            if not data:
                env["ir.model.data"].create({"module": module, "name": xml_name, "model": "account.account", "res_id": account.id, "noupdate": True})
            else:
                data.write({"noupdate": True, "res_id": account.id})
        else:
            if not values.get("account_type"):
                values["account_type"] = "asset_current"
            account = _get_or_create(env, "account.account", row["id"], values)

        if row["opening_debit"] or row["opening_credit"]:
            opening_totals[account] = [row["opening_debit"] or 0.0, row["opening_credit"] or 0.0]

    _write_opening_balances(env, opening_totals)


def _import_account_journal(env, wb):
    columns = ("id", "sequence", "name", "type", "code", "default_account_id", "Bank Feed")
    company = env.ref("base.main_company")
    for row in _sheet_rows(wb, "account.journal", columns):
        values = {
            "sequence": row["sequence"],
            "name": row["name"],
            "type": JOURNAL_TYPES[row["type"]],
            "code": row["code"],
            "bank_statements_source": row["Bank Feed"],
        }
        account = _account_by_code_name(env, row["default_account_id"])
        if account:
            values["default_account_id"] = account.id

        # Search by code within company first to respect UNIQUE(company_id, code) constraint
        journal = False
        if row.get("code"):
            journal = env["account.journal"].with_context(active_test=False).search([
                ("code", "=", str(row["code"]).strip()),
                ("company_id", "=", company.id),
            ], limit=1)
        if not journal:
            journal = env.ref(row["id"], raise_if_not_found=False)
        if not journal and row.get("name"):
            journal = env["account.journal"].with_context(active_test=False).search([
                ("name", "=", str(row["name"]).strip()),
                ("company_id", "=", company.id),
            ], limit=1)

        if journal:
            journal.write(values)
            if "." in row["id"]:
                module, xml_name = row["id"].split(".", 1)
            else:
                module, xml_name = MODULE, row["id"]
            data = env["ir.model.data"].search([("module", "=", module), ("name", "=", xml_name)])
            if not data:
                env["ir.model.data"].create({"module": module, "name": xml_name, "model": "account.journal", "res_id": journal.id, "noupdate": True})
            else:
                data.write({"noupdate": True, "res_id": journal.id})
        else:
            _get_or_create(env, "account.journal", row["id"], values)


def _import_account_asset(env, wb, balance_date):
    columns = (
        "id", "name", "acquisition_date", "original_value", "already_depreciated_amount_import",
        "account_asset_id", "account_depreciation_id", "account_depreciation_expense_id",
        "method", "method_number", "method_period", "Jurnal",
    )
    opening_totals = {}  # account.account recordset -> [opening_debit, opening_credit]

    def _add_opening(account, debit=0.0, credit=0.0):
        if account:
            totals = opening_totals.setdefault(account, [0.0, 0.0])
            totals[0] += debit
            totals[1] += credit

    for row in _sheet_rows(wb, "account.asset", columns):
        journal = env["account.journal"].search([("name", "=", row["Jurnal"])], limit=1)
        asset_account = _account_by_code_name(env, row["account_asset_id"])
        depreciation_account = _account_by_code_name(env, row["account_depreciation_id"])
        method_period = ASSET_PERIODS[row["method_period"]]
        acquisition_date = _parse_sheet_date(row["acquisition_date"], "acquisition_date")

        already_depreciated = row["already_depreciated_amount_import"]
        if already_depreciated in (None, ""):
            elapsed = _elapsed_depreciation_periods(acquisition_date, balance_date, method_period)
            elapsed = min(elapsed, row["method_number"])
            already_depreciated = (row["original_value"] or 0.0) / row["method_number"] * elapsed

        values = {
            "name": row["name"],
            "acquisition_date": acquisition_date,
            "original_value": row["original_value"],
            "already_depreciated_amount_import": already_depreciated,
            "account_asset_id": asset_account.id,
            "account_depreciation_id": depreciation_account.id,
            "account_depreciation_expense_id": _account_by_code_name(env, row["account_depreciation_expense_id"]).id,
            "method": ASSET_METHODS[row["method"]],
            "method_number": row["method_number"],
            "method_period": method_period,
            "journal_id": journal.id,
        }
        asset = _get_or_create(env, "account.asset", row["id"], values)
        if VALIDATE_IMPORTED_ASSETS and asset.state == "draft":
            asset.validate()

        _add_opening(asset_account, debit=row["original_value"] or 0.0)
        _add_opening(depreciation_account, credit=already_depreciated)

    _write_opening_balances(env, opening_totals)


def _reconcile_liquidity_transfer(env):
    """Auto-reconcile kas/bank opening entries against their Liquidity
    Transfer counterpart once both sides are posted (e.g. after the
    account.account opening_debit move has been reviewed and posted).
    """
    company = env.ref("base.main_company")
    lines = env["account.move.line"].search([
        ("account_id", "=", company.transfer_account_id.id),
        ("reconciled", "=", False),
        ("move_id.state", "=", "posted"),
    ])
    if lines:
        lines.reconcile()


def _check_user_journal_entries(env, company):
    """Cek apakah sudah ada journal entries operasional yang diinput oleh user.
    Mengembalikan recordset account.move buatan user jika ditemukan.
    """
    opening_move_id = company.account_opening_move_id.id if company.account_opening_move_id else None

    domain = [("company_id", "=", company.id)]
    moves = env["account.move"].search(domain)

    user_moves = env["account.move"]
    for m in moves:
        # Lewati opening journal entry neraca bawaan/sistem
        if opening_move_id and m.id == opening_move_id:
            continue

        # Lewati opening bank statement lines dari import saldo awal
        if m.statement_line_id:
            ref = (m.statement_line_id.payment_ref or "").lower()
            if "saldo awal" in ref or "opening" in ref:
                continue

        # Lewati opening vendor_bill / customer_invoice yang dibuat oleh import ini
        imd = env["ir.model.data"].search([
            ("model", "=", "account.move"),
            ("res_id", "=", m.id),
            ("module", "in", (MODULE, "__import__")),
        ], limit=1)
        if imd and ("vendor_bill" in imd.name or "customer_invoice" in imd.name):
            continue

        user_moves |= m

    return user_moves


def _cleanup_previous_data(env, wb, company):
    """Membersihkan sisa data master dan saldo awal dari import sebelumnya
    atau sisa kloning database ketika belum ada transaksi operasional user.
    """
    _logger.info("Membersihkan data sisa import sebelumnya untuk company %s...", company.name)

    # 0. Bersihkan mapping ir.model.data dinamis lama agar ID baru tidak tertukar/overwrite
    imds = env["ir.model.data"].search([
        ("module", "=", MODULE),
        ("model", "in", ("account.account", "account.journal", "account.asset", "account.bank.statement.line")),
    ])
    if imds:
        imds.unlink()

    # 1. Bersihkan statement lines saldo awal lama
    st_lines = env["account.bank.statement.line"].search([("company_id", "=", company.id)])
    for st in st_lines:
        if st.is_reconciled:
            try:
                st.line_ids.remove_move_reconcile()
            except Exception:
                pass
    if st_lines:
        st_lines.unlink()

    # 2. Reset opening balance move neraca lama
    if company.account_opening_move_id:
        op_move = company.account_opening_move_id
        if op_move.state == "posted":
            op_move.button_draft()
        op_move.line_ids.unlink()

    # 3. Bersihkan jurnal custom lama yang TIDAK ADA di spreadsheet baru
    sheet_journal_codes = set()
    sheet_journal_names = set()
    if "account.journal" in wb.sheetnames:
        for r in _sheet_rows(wb, "account.journal", ("id", "code", "name")):
            if r.get("code"):
                sheet_journal_codes.add(str(r["code"]).strip())
            if r.get("name"):
                sheet_journal_names.add(str(r["name"]).strip())

    standard_journal_codes = {"INV", "BILL", "MISC", "EXCH", "CABA", "TAX"}
    leftover_journals = env["account.journal"].search([
        ("company_id", "=", company.id),
        ("code", "not in", list(standard_journal_codes | sheet_journal_codes)),
    ])
    for j in list(leftover_journals):
        if j.name in sheet_journal_names:
            continue
        # JANGAN PERNAH hapus jurnal bawaan Odoo (memiliki External ID selain modul import ini atau id <= 8)
        is_system_journal = env["ir.model.data"].search_count([
            ("model", "=", "account.journal"),
            ("res_id", "=", j.id),
            ("module", "not in", (MODULE, "__import__")),
        ])
        if is_system_journal or j.id <= 8:
            continue

        try:
            with env.cr.savepoint():
                imds = env["ir.model.data"].search([("model", "=", "account.journal"), ("res_id", "=", j.id)])
                imds.unlink()
                j.unlink()
        except Exception:
            j.active = False

    # 4. Bersihkan akun COA custom lama yang TIDAK ADA di spreadsheet baru
    sheet_account_codes = set()
    if "account.account" in wb.sheetnames:
        for r in _sheet_rows(wb, "account.account", ("code",)):
            if r.get("code"):
                sheet_account_codes.add(str(r["code"]).strip())

    leftover_accounts = env["account.account"].with_context(active_test=False).search([
        ("company_ids", "in", company.id),
        ("code", "not in", list(sheet_account_codes)),
    ])
    journal_account_ids = set(
        env["account.journal"].search([("company_id", "=", company.id)]).mapped("default_account_id.id")
    ) | set(
        env["account.journal"].search([("company_id", "=", company.id)]).mapped("suspense_account_id.id")
    )

    for acc in leftover_accounts:
        # JANGAN PERNAH hapus akun bawaan Odoo (memiliki External ID dari modul sistem, misal 'account', 'l10n_id', 'base')
        is_system_account = env["ir.model.data"].search_count([
            ("model", "=", "account.account"),
            ("res_id", "=", acc.id),
            ("module", "not in", (MODULE, "__import__")),
        ])
        if is_system_account:
            # Akun bawaan sistem Odoo: jangan dihapus
            continue

        # Jangan hapus akun jika dipakai di konfigurasi perusahaan
        if acc.id in (
            company.transfer_account_id.id if company.transfer_account_id else 0,
            company.account_journal_suspense_account_id.id if company.account_journal_suspense_account_id else 0,
        ) or acc.id in journal_account_ids:
            acc.active = False
            continue

        try:
            with env.cr.savepoint():
                imds = env["ir.model.data"].search([("model", "=", "account.account"), ("res_id", "=", acc.id)])
                imds.unlink()
                acc.unlink()
        except Exception:
            acc.active = False


def run_import(env, wb):
    """Run the full master-data import against an already-open workbook.
    Shared entry point for import_bundled_data() (a client module's
    post_init_hook) and the Settings > Import Data Master wizard (a
    user-uploaded file).
    """
    company = env.ref("base.main_company")

    # Validasi: Jika sudah ada transaksi/journal entry operasional yang diinput user, tolak import
    user_moves = _check_user_journal_entries(env, company)
    if user_moves:
        sample_names = ", ".join([str(m.name or m.ref or f"Draft #{m.id}") for m in user_moves[:5]])
        raise UserError(
            f"Import Ditolak: Ditemukan {len(user_moves)} transaksi/journal entry operasional "
            f"yang telah diinput oleh user ({sample_names}).\n\n"
            f"Untuk menjaga integritas dan validitas data akuntansi, proses import data master "
            f"hanya dapat dijalankan jika belum ada transaksi operasional yang dibuat manual.\n"
            f"Harap batalkan atau hapus journal entries tersebut terlebih dahulu jika Anda ingin "
            f"mengulang import data master dan saldo awal."
        )

    # Bersihkan sisa data lama dari import sebelumnya atau database hasil clone
    _cleanup_previous_data(env, wb, company)

    balance_date = _read_opening_balance_date(wb)
    # account_opening_date perlu tanggal *awal* tahun buku, bukan tanggal
    # cutover-nya - Odoo men-tanggal-kan opening move sehari sebelum awal
    # tahun buku, jadi arahnya kebalik dari balance_date.
    fiscal_year_start = balance_date + timedelta(days=1)
    if fiscal_year_start.day != 1:
        _logger.warning(
            'OPENING_BALANCE_DATE (%s) menghasilkan awal tahun buku %s, bukan tanggal 1 - '
            "cek lagi kalau ini bukan disengaja (mis. salah pilih tanggal).",
            balance_date, fiscal_year_start,
        )

    company.write({"account_opening_date": fiscal_year_start})
    if company.account_opening_move_id and company.account_opening_move_id.state == "posted":
        company.account_opening_move_id.button_draft()
    if company.account_opening_move_id and company.account_opening_move_id.state == "draft":
        company.account_opening_move_id.date = balance_date

    _import_company(env, wb)
    _import_res_partner(env, wb)
    _import_account_account(env, wb)
    _import_account_journal(env, wb)
    _import_account_asset(env, wb, balance_date)
    _import_kas_bank(env, wb, balance_date)
    _import_opening_move(env, wb, "vendor_bill", "supplier_rank", balance_date)
    _import_opening_move(env, wb, "customer_invoice", "customer_rank", balance_date)
    _reconcile_liquidity_transfer(env)
    _import_account_analytic_plan(env, wb)
    _import_account_analytic_account(env, wb)
    _import_account_report(env, wb)
    _import_account_report_line(env, wb)
